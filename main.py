import os
from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook
from fastapi.middleware.cors import CORSMiddleware

from typing import List, Dict, Any
from pathlib import Path


app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Serve the HTML frontend
@app.get("/", response_class=HTMLResponse)
async def serve_frontend_of_proj():
    html = Path("project1.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html, media_type="text/html")


# Serve the HTML frontend
@app.get("/project1", response_class=HTMLResponse)
async def serve_frontend_of_proj_1():
    html = Path("project1.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html, media_type="text/html")


# Serve the HTML frontend
@app.get("/project2", response_class=HTMLResponse)
async def serve_frontend_of_proj_2():
    html = Path("project2.html").read_text(encoding="utf-8")
    return HTMLResponse(content=html, media_type="text/html")


class Record(BaseModel):
    sr_no: int = Field(..., alias="Sr No")
    segment: str = Field(..., alias="Segment")
    sub_segment: str = Field(..., alias="Sub Segment")
    action_pointers: str = Field(..., alias="Action Pointers")
    timeline: str = Field(..., alias="Timeline")
    status: str = Field(..., alias="Status")
    # actions: str      = Field(..., alias="Actions")

    class Config:
        # allow_population_by_field_name = True
        validate_by_name = True
        allow_population_by_alias = True


@app.post("/{project_id}/append-record/")
async def append_record(project_id: str, record: Record):
    if project_id != "project_1":
        file_path = "project_1_records.xlsx"
    elif project_id != "project_2":
        file_path = "project_2_records.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid project ID.")

    data = record.model_dump(by_alias=True)

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.append(list(data.keys()))
    else:
        wb = load_workbook(file_path)
        ws = wb.active

    ws.append(list(data.values()))
    wb.save(file_path)

    return {"message": "Record appended successfully."}


@app.get("/{project_id}/records/", response_model=List[Record])
async def get_records(project_id: str):
    if project_id != "project_1":
        file_path = "project_1_records.xlsx"
    elif project_id != "project_2":
        file_path = "project_2_records.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid project ID.")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Records file not found.")

    wb = load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return []

    headers = rows[0]
    records: List[Dict[str, Any]] = []
    for row in rows[1:]:
        records.append(dict(zip(headers, row)))

    # FastAPI will use the Record model to validate and serialize
    return records


@app.delete("/{project_id}/record/{sr_no}")
async def delete_record(project_id: str, sr_no: int):
    if project_id != "project_1":
        file_path = "project_1_records.xlsx"
    elif project_id != "project_2":
        file_path = "project_2_records.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid project ID.")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Records file not found.")

    wb = load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="No data available.")

    headers = rows[0]
    if "Sr No" not in headers:
        raise HTTPException(status_code=400, detail="Sr No column missing.")
    sr_index = headers.index("Sr No")

    row_to_delete = None
    for idx_row, row in enumerate(rows[1:], start=2):
        if row[sr_index] == sr_no:
            row_to_delete = idx_row
            break
    if row_to_delete is None:
        raise HTTPException(status_code=404, detail="Record not found.")

    ws.delete_rows(row_to_delete)
    wb.save(file_path)
    return {"message": f"Record with Sr No {sr_no} deleted successfully."}


# Update a record by Sr No
@app.put("/{project_id}/record/{sr_no}")
async def update_record(project_id: str, sr_no: int, record: Record):
    if project_id != "project_1":
        file_path = "project_1_records.xlsx"
    elif project_id != "project_2":
        file_path = "project_2_records.xlsx"
    else:
        raise HTTPException(status_code=400, detail="Invalid project ID.")

    if not os.path.exists(file_path):
        raise HTTPException(status_code=404, detail="Records file not found.")
    wb = load_workbook(file_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="No data available.")
    headers = rows[0]
    # Find index for Sr No and row
    if "Sr No" not in headers:
        raise HTTPException(status_code=400, detail="Sr No column missing.")
    sr_index = headers.index("Sr No")
    row_to_update = None
    for idx_row, row in enumerate(rows[1:], start=2):
        if row[sr_index] == sr_no:
            row_to_update = idx_row
            break
    if row_to_update is None:
        raise HTTPException(status_code=404, detail="Record not found.")
    # Prepare new data
    data = record.dict(by_alias=True)
    # Update cells
    for col_index, header in enumerate(headers, start=1):
        ws.cell(row=row_to_update, column=col_index, value=data.get(header))
    wb.save(file_path)
    return {"message": f"Record with Sr No {sr_no} updated successfully."}


from fastapi.responses import StreamingResponse
import asyncio


async def word_stream(docsString):
    for word in docsString.split():
        yield word + " "
        await asyncio.sleep(1)


class Test(BaseModel):
    docsString: str = ""


@app.get("/teststreamapi")
async def stream_api(docsString: Test):
    return StreamingResponse(word_stream(docsString), media_type="text/plain")
